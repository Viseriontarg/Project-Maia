# report_generator.py

import pandas as pd
import numpy as np
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
import os

# --- Predefined Named Styles ---
currency_style = NamedStyle(name='currency', number_format='$#,##0.00')
integer_style = NamedStyle(name='integer', number_format='#,##0')
days_style = NamedStyle(name='days', number_format='#,##0" days"')
percent_style = NamedStyle(name='percent_decimal', number_format='0.00%')

def _add_named_styles(workbook):
    for style in [currency_style, integer_style, days_style, percent_style]:
        if style.name not in workbook.style_names:
            workbook.add_named_style(style)

def _get_formatted_value(data_dict, key, default_val='N/A'):
    if not isinstance(data_dict, dict): return default_val
    val = data_dict.get(key)
    return val if val is not None else default_val

def convert_numpy_types(data):
    if isinstance(data, dict): return {k: convert_numpy_types(v) for k, v in data.items()}
    if isinstance(data, list): return [convert_numpy_types(i) for i in data]
    if isinstance(data, np.floating): return float(data)
    if isinstance(data, np.integer): return int(data)
    if isinstance(data, np.ndarray): return data.tolist()
    return data

def _style_dataframe_output(ws, df_num_rows, df_num_cols, start_row=1, start_col=1, col_styles=None, auto_width_padding=3, workbook=None):
    header_font = Font(bold=True, name='Calibri', size=11)
    thin_border_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    for c_offset in range(df_num_cols):
        cell = ws.cell(row=start_row, column=start_col + c_offset)
        cell.font = header_font; cell.border = cell_border; cell.alignment = Alignment(horizontal='left', vertical='center')
    for r_offset in range(df_num_rows):
        for c_offset in range(df_num_cols):
            cell = ws.cell(row=start_row + 1 + r_offset, column=start_col + c_offset)
            cell.border = cell_border
            if col_styles and workbook:
                # Use column index (0-based) for style lookup
                # This check is safer than direct key access
                if c_offset in col_styles:
                     cell.style = col_styles[c_offset]

    for c_offset in range(df_num_cols):
        column_letter = get_column_letter(start_col + c_offset)
        # Simplified width calculation
        max_len = max((len(str(ws.cell(row=r, column=start_col + c_offset).value or "")) for r in range(start_row, start_row + df_num_rows + 2)), default=10)
        ws.column_dimensions[column_letter].width = max_len + auto_width_padding

# --- Sheet Creation Functions ---
def _create_summary_sheet(writer, input_data, simulation_results):
    ws_name = "Summary"
    ws = writer.sheets[ws_name]
    wb = writer.book

    summary_items = [
        ("Project Name", _get_formatted_value(input_data, "projectName")),
        ("User Target Base Cost", _get_formatted_value(input_data, "targetBaseCost")),
        ("User Target Base Duration", _get_formatted_value(input_data, "targetBaseDuration")),
        ("Team Size", _get_formatted_value(input_data, "teamSize")),
        ("Daily Cost", _get_formatted_value(input_data, "dailyCost")),
    ]
    
    duration_analysis = simulation_results.get("durationAnalysis", {})
    cost_analysis = simulation_results.get("costAnalysis", {})

    summary_items.extend([
        ("P10 (Optimistic) Duration", _get_formatted_value(duration_analysis, "p10")),
        ("P10 (Optimistic) Cost", _get_formatted_value(cost_analysis, "p10")),
        ("P50 (Likely) Duration", _get_formatted_value(duration_analysis, "p50")),
        ("P50 (Likely) Cost", _get_formatted_value(cost_analysis, "p50")),
        ("P90 (Pessimistic) Duration", _get_formatted_value(duration_analysis, "p90")),
        ("P90 (Pessimistic) Cost", _get_formatted_value(cost_analysis, "p90")),
        ("P95 (Worst Case) Duration", _get_formatted_value(duration_analysis, "p95")),
        ("P95 (Worst Case) Cost", _get_formatted_value(cost_analysis, "p95")),
    ])
    
    top_risks_by_days_data = simulation_results.get('topRiskDriversByDays', [])
    for i in range(3):
        metric_name = f"Top Risk {i+1} - Total Aggregated Days Impact (All Runs)"
        if i < len(top_risks_by_days_data):
            risk = top_risks_by_days_data[i]
            name = risk.get('name', 'Unknown')
            impact = risk.get('totalImpactDays', 0.0)
            summary_items.append((metric_name, f"{name} ({impact:.1f} days)"))
        else: summary_items.append((metric_name, "N/A"))

    top_risks_by_cost_data = simulation_results.get('topRiskDriversByCost', [])
    for i in range(3):
        metric_name = f"Top Risk {i+1} - Total Aggregated Cost Impact (All Runs)"
        if i < len(top_risks_by_cost_data):
            risk = top_risks_by_cost_data[i]
            name = risk.get('name', 'Unknown')
            impact = risk.get('totalImpactCost', 0.0)
            summary_items.append((metric_name, f"{name} (${impact:,.0f})"))
        else: summary_items.append((metric_name, "N/A"))

    header_font = Font(bold=True, name='Calibri', size=11)
    thin_border_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    ws['A1'] = 'Metric'; ws['B1'] = 'Value'
    ws['A1'].font = header_font; ws['A1'].border = cell_border
    ws['B1'].font = header_font; ws['B1'].border = cell_border
    
    for idx, (metric, value) in enumerate(summary_items, start=2):
        cell_a = ws[f'A{idx}']; cell_a.value = metric
        cell_b = ws[f'B{idx}']; cell_b.value = value
        cell_a.font = header_font; cell_a.border = cell_border
        cell_b.border = cell_border

        if isinstance(value, (int, float)):
            if "Cost" in metric: cell_b.style = currency_style.name
            elif "Duration" in metric: cell_b.style = days_style.name
            elif "Team Size" in metric: cell_b.style = integer_style.name
            else: cell_b.style = integer_style.name
    
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 30

    distribution_sample = simulation_results.get('distributionSample', {})
    durations_sample = distribution_sample.get('durations', [])
    if durations_sample and len(durations_sample) > 1:
        num_bins = 10
        counts, bin_edges = np.histogram(durations_sample, bins=num_bins)
        chart_data_start_row = 2; chart_data_start_col = 20
        ws.cell(row=chart_data_start_row - 1, column=chart_data_start_col).value = "Chart Freq"
        ws.cell(row=chart_data_start_row - 1, column=chart_data_start_col + 1).value = "Chart Bins"
        bin_labels = [f"{bin_edges[i]:.0f}-{bin_edges[i+1]:.0f}" for i in range(len(bin_edges)-1)]
        for i, (count, label) in enumerate(zip(counts, bin_labels)):
            ws.cell(row=chart_data_start_row + i, column=chart_data_start_col).value = count
            ws.cell(row=chart_data_start_row + i, column=chart_data_start_col + 1).value = label
            
        chart = BarChart(); chart.type = "col"; chart.style = 10; chart.grouping = "standard"
        chart.title = "Duration Distribution"; chart.y_axis.title = "Frequency"; chart.x_axis.title = "Duration (Days)"
        data_ref = Reference(ws, min_col=chart_data_start_col, min_row=chart_data_start_row, max_row=chart_data_start_row + len(counts) - 1)
        cat_ref = Reference(ws, min_col=chart_data_start_col + 1, min_row=chart_data_start_row, max_row=chart_data_start_row + len(bin_labels) - 1)
        series = Series(data_ref, title="Frequency"); series.graphicalProperties = GraphicalProperties(solidFill="5B9BD5")
        chart.series.append(series); chart.set_categories(cat_ref); chart.legend = None; chart.y_axis.majorGridlines = None
        ws.add_chart(chart, "D2"); chart.width = 15; chart.height = 7.5

def _create_inputs_assumptions_sheet(writer, input_data):
    ws_name = "Inputs & Assumptions"
    wb = writer.book
    ws = wb.create_sheet(title=ws_name)
    writer.sheets[ws_name] = ws
    current_row_excel = 1
    title_font = Font(bold=True, size=14)

    # --- DYNAMIC TASK HANDLING ---
    tasks_for_df = []
    if input_data.get('phases'):
        for phase in input_data['phases']:
            for task in phase.get('tasks', []):
                task_row = {
                    'Phase': _get_formatted_value(phase, "phaseName"),
                    'Task Name': _get_formatted_value(task, "name"),
                }
                # Merge duration parameters directly into the row
                if isinstance(task.get('duration_params'), dict):
                    task_row.update(task['duration_params'])
                tasks_for_df.append(task_row)

    if tasks_for_df:
        # Let Pandas create the DataFrame, handling all columns dynamically
        df_tasks = pd.DataFrame(tasks_for_df).fillna('N/A')
        # Standardize column naming for presentation
        df_tasks.rename(columns={
            'type': 'Distribution Type', 'optimistic': 'Optimistic (days)',
            'most_likely': 'Most Likely (days)', 'pessimistic': 'Pessimistic (days)',
            'mean': 'Mean (days)', 'std_dev': 'Std Dev (days)',
            'min': 'Min (days)', 'max': 'Max (days)'
        }, inplace=True)

        ws.cell(row=current_row_excel, column=1, value="Project Tasks (from Input Data)").font = title_font
        current_row_excel += 1
        df_tasks.to_excel(writer, sheet_name=ws_name, startrow=current_row_excel-1, index=False)
        
        # DYNAMIC STYLING
        task_col_styles = {}
        for i, col_name in enumerate(df_tasks.columns):
            if '(days)' in col_name: task_col_styles[i] = integer_style.name
        
        _style_dataframe_output(ws, len(df_tasks), len(df_tasks.columns), start_row=current_row_excel, workbook=wb, col_styles=task_col_styles)
        current_row_excel += len(df_tasks) + 3

    # --- DYNAMIC RISK HANDLING ---
    risks_for_df = []
    if input_data.get('risks'):
        for risk in input_data['risks']:
            risk_row = {
                'Risk Name': _get_formatted_value(risk, "name"),
                'Probability': _get_formatted_value(risk, "probability"),
            }
            # Unpack schedule and cost impacts with prefixes
            if isinstance(risk.get('impactDays'), dict):
                for k, v in risk['impactDays'].items():
                    risk_row[f'Sched. Impact {k}'] = v
            if isinstance(risk.get('impactCost'), dict):
                for k, v in risk['impactCost'].items():
                    risk_row[f'Cost Impact {k}'] = v
            risks_for_df.append(risk_row)

    if risks_for_df:
        # Let Pandas handle all columns dynamically
        df_risks = pd.DataFrame(risks_for_df).fillna('N/A')
        # Standardize column naming for presentation
        df_risks.rename(columns={
            'Sched. Impact type': 'Sched. Impact Dist. Type', 'Sched. Impact optimistic': 'Optimistic Sched. Impact (days)',
            'Sched. Impact most_likely': 'Most Likely Sched. Impact (days)', 'Sched. Impact pessimistic': 'Pessimistic Sched. Impact (days)',
            'Sched. Impact mean': 'Mean Sched. Impact (days)', 'Sched. Impact std_dev': 'Std Dev Sched. Impact (days)',
            'Sched. Impact min': 'Min Sched. Impact (days)', 'Sched. Impact max': 'Max Sched. Impact (days)',
            'Cost Impact type': 'Cost Impact Dist. Type', 'Cost Impact optimistic': 'Optimistic Cost Impact ($)',
            'Cost Impact most_likely': 'Most Likely Cost Impact ($)', 'Cost Impact pessimistic': 'Pessimistic Cost Impact ($)',
            'Cost Impact mean': 'Mean Cost Impact ($)', 'Cost Impact std_dev': 'Std Dev Cost Impact ($)',
            'Cost Impact min': 'Min Cost Impact ($)', 'Cost Impact max': 'Max Cost Impact ($)',
        }, inplace=True)
        
        ws.cell(row=current_row_excel, column=1, value="Input Risk Assumptions").font = title_font
        current_row_excel += 1
        df_risks.to_excel(writer, sheet_name=ws_name, startrow=current_row_excel-1, index=False)

        # DYNAMIC STYLING
        risk_col_styles = {}
        for i, col_name in enumerate(df_risks.columns):
            if 'Probability' in col_name: risk_col_styles[i] = percent_style.name
            elif '(days)' in col_name: risk_col_styles[i] = integer_style.name
            elif '($)' in col_name: risk_col_styles[i] = currency_style.name

        _style_dataframe_output(ws, len(df_risks), len(df_risks.columns), start_row=current_row_excel, workbook=wb, col_styles=risk_col_styles, auto_width_padding=2)

def _create_distribution_data_sheet(writer, simulation_results):
    ws_name = "Distribution Data"
    wb = writer.book
    ws = wb.create_sheet(ws_name)
    writer.sheets[ws_name] = ws
    distribution_sample = simulation_results.get('distributionSample', {})
    durations = distribution_sample.get('durations', [])
    costs = distribution_sample.get('costs', [])
    if durations:
        df_dist = pd.DataFrame({'Sample Durations': durations, 'Sample Costs': costs})
        df_dist.to_excel(writer, sheet_name=ws_name, index=False)
        col_styles = {0: integer_style.name, 1: currency_style.name}
        _style_dataframe_output(ws, len(df_dist), 2, start_row=1, col_styles=col_styles, workbook=wb)

# --- Main Report Generation Function ---
def create_excel_report(input_data, simulation_results):
    project_name_sanitized = "Simulation_Report"
    if input_data and isinstance(input_data, dict) and input_data.get("projectName"):
        project_name_sanitized = "".join(c for c in input_data["projectName"] if c.isalnum() or c in (' ', '_')).rstrip().replace(' ', '_')
    
    reports_dir = "generated_reports"
    if not os.path.exists(reports_dir): os.makedirs(reports_dir)
    file_path = os.path.join(reports_dir, f"{project_name_sanitized}.xlsx")

    try:
        processed_simulation_results = convert_numpy_types(simulation_results)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            wb = writer.book
            _add_named_styles(wb)
            
            summary_ws = wb.create_sheet("Summary", 0)
            writer.sheets["Summary"] = summary_ws
            
            _create_summary_sheet(writer, input_data, processed_simulation_results)
            _create_inputs_assumptions_sheet(writer, input_data)
            _create_distribution_data_sheet(writer, processed_simulation_results)

            if "Sheet" in wb.sheetnames: del wb["Sheet"]
            wb.active = wb.sheetnames.index("Summary")
            
        print(f"Excel report successfully generated: {file_path}")
        return file_path
    except Exception as e:
        print(f"Error creating Excel report at {file_path}: {e}")
        import traceback; traceback.print_exc()
        return None

if __name__ == '__main__':
    # This block allows you to test the report generation directly
    # by creating mock data without running the full simulation.
    print("--- Running Report Generator Test ---")

    # 1. Create Mock Input Data (what ai_breakdown.py would produce)
    # This data is intentionally "sparse" to test the dynamic columns.
    mock_input_data = {
        "projectName": "Mock Project for Testing",
        "targetBaseCost": 50000,
        "targetBaseDuration": 60,
        "teamSize": 5,
        "dailyCost": 650,
        "phases": [
            {
                "phaseName": "Phase 1: Design",
                "tasks": [
                    {
                        "name": "Create Mockups",
                        "duration_params": {"type": "PERT", "optimistic": 8, "most_likely": 10, "pessimistic": 15}
                    },
                    {
                        "name": "Technical Specification",
                        "duration_params": {"type": "Normal", "mean": 12, "std_dev": 2}
                    }
                ]
            }
        ],
        "risks": [
            {
                "name": "Key Designer Unavailable",
                "probability": 0.15,
                "impactDays": {"type": "Uniform", "min": 5, "max": 10},
                "impactCost": {"type": "Triangular", "optimistic": 2000, "most_likely": 3000, "pessimistic": 5000}
            },
            {
                "name": "Software Licensing Delays",
                "probability": 0.20,
                "impactDays": {"type": "Normal", "mean": 7, "std_dev": 1}
                # This risk has no direct cost impact, testing missing keys
            }
        ]
    }
    print("[Step 1] Created mock input data.")

    # 2. Create Mock Simulation Results (what a simulation would produce)
    np.random.seed(42) # for reproducibility
    mock_sim_results = {
        "durationAnalysis": {"p10": 65, "p50": 72, "p90": 85, "p95": 92},
        "costAnalysis": {"p10": 55000, "p50": 61000, "p90": 78000, "p95": 85000},
        "topRiskDriversByDays": [
            {"name": "Key Designer Unavailable", "totalImpactDays": 150.5},
            {"name": "Software Licensing Delays", "totalImpactDays": 120.2}
        ],
        "topRiskDriversByCost": [
            {"name": "Key Designer Unavailable", "totalImpactCost": 45000}
        ],
        "distributionSample": {
            "durations": np.random.normal(72, 8, 10000),
            "costs": np.random.normal(61000, 7000, 10000)
        }
    }
    print("[Step 2] Created mock simulation results.")

    # 3. Call the main report generation function
    print("[Step 3] Generating Excel report...")
    file_path = create_excel_report(mock_input_data, mock_sim_results)

    if file_path:
        print(f"\n--- Report Generator Test Complete ---")
        print(f"Test report saved to: {os.path.abspath(file_path)}")
    else:
        print(f"\n--- Report Generator Test Failed ---")